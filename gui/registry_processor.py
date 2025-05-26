import pandas as pd
import time
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
import openpyxl
from selenium.common.exceptions import TimeoutException
from datetime import datetime
import sys
import platform
from webdriver_manager.chrome import ChromeDriverManager

class RegistryProcessor:
    def __init__(self, progress_callback=None, status_callback=None):
        """
        Initialize the registry processor
        
        Args:
            progress_callback: Function to call with progress updates (current, total, name)
            status_callback: Function to call with status messages
        """
        self.progress_callback = progress_callback
        self.status_callback = status_callback
        self.driver = None
        self.should_stop = False
        self.disclaimer_accepted = False  # Track if disclaimer has been accepted
        
    def _log_status(self, message):
        """Log status message to callback if available"""
        if self.status_callback:
            self.status_callback(message)
        print(message)  # Also print for debugging
    
    def _update_progress(self, current, total, name=""):
        """Update progress via callback if available"""
        if self.progress_callback:
            self.progress_callback(current, total, name)
    
    def setup_driver(self):
        """Initialize the headless Chrome driver"""
        try:
            self._log_status("Initializing browser...")
            chrome_options = Options()
            # Disable location requests
            prefs = {"profile.default_content_setting_values.geolocation": 2}
            chrome_options.add_experimental_option("prefs", prefs)
            # Headless mode configuration
            chrome_options.add_argument('--headless=new')
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_argument('--window-size=1920,1080')
            chrome_options.add_argument('--disable-web-security')
            chrome_options.add_argument('--disable-features=VizDisplayCompositor')
            # Use webdriver-manager to handle ChromeDriver
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            self._log_status("Browser initialized successfully")
            return True
        except Exception as e:
            self._log_status(f"Error initializing browser: {str(e)}")
            return False
    
    def read_excel_file(self, file_path):
        """
        Read names from Excel file
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            pandas.DataFrame with columns: First Name, Last Name, Gender
        """
        try:
            self._log_status(f"Reading Excel file: {os.path.basename(file_path)}")
            
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            rows = []
            row_num = 4  # Start from row 4 (assuming headers in row 3)
            
            while True:
                first_name = ws[f'B{row_num}'].value
                last_name = ws[f'D{row_num}'].value
                gender = ws[f'G{row_num}'].value
                
                # Stop if we hit empty rows
                if not first_name and not last_name:
                    break
                
                # Only add if we have both first and last name
                if first_name and last_name:
                    gender_value = gender.lower() if gender else 'unknown'
                    rows.append({
                        'First Name': str(first_name).strip(),
                        'Last Name': str(last_name).strip(),
                        'Gender': gender_value
                    })
                
                row_num += 1
            
            df = pd.DataFrame(rows)
            self._log_status(f"Found {len(df)} valid names to process")
            return df
            
        except Exception as e:
            self._log_status(f"Error reading Excel file: {str(e)}")
            return None
    
    def search_single_name(self, first_name, last_name, gender):
        """
        Search a single name in the registry
        
        Returns:
            str: 'no results found', 'warning', or error message
        """
        url = 'https://sexoffenderregistry.ar.gov/public/#/'
        
        try:
            self.driver.get(url)
            time.sleep(2)
            
            # Handle disclaimer if present, only on first search
            if not self.disclaimer_accepted:
                try:
                    agree_button = WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, "//button[contains(.,'I Agree to the Disclaimer Above')]") )
                    )
                    # Scroll to the very bottom of the page
                    self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(0.7)
                    # Force click the button with JavaScript
                    self.driver.execute_script("arguments[0].click();", agree_button)
                    time.sleep(1)
                    self.disclaimer_accepted = True
                except Exception as e:
                    self._log_status(f"Disclaimer button not found or not clickable: {e}")
                    # Even if not found, set as accepted to avoid repeated attempts
                    self.disclaimer_accepted = True
            
            # Click Name Search tab
            WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//*[contains(text(),'Name Search')]"))
            ).click()
            time.sleep(2)
            
            # Fill in form fields
            first_name_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//input[@id='firstName']"))
            )
            last_name_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//input[@id='lastName']"))
            )
            
            first_name_field.clear()
            first_name_field.send_keys(first_name)
            first_name_field.send_keys(Keys.TAB)
            time.sleep(0.5)
            
            last_name_field.clear()
            last_name_field.send_keys(last_name)
            last_name_field.send_keys(Keys.TAB)
            time.sleep(0.5)
            
            # Set gender
            gender_select = Select(self.driver.find_element(By.XPATH, "//select[@id='gender']"))
            gender_select.select_by_value(gender)
            
            # Click search
            search_button = self.driver.find_element(By.XPATH, "//button[@type='submit' and contains(.,'Search')]")
            search_button.click()
            time.sleep(2)
            
            # Check results
            try:
                WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//div[contains(@class,'no-result') or contains(@class,'col-12 col-md-auto text-center text-start-md')]"))
                )
                
                # Check for no results
                try:
                    self.driver.find_element(By.XPATH, "//div[contains(@class,'no-result') and contains(.,'No results')]")
                    return 'no results found'
                except Exception:
                    # Check for results card
                    try:
                        self.driver.find_element(By.XPATH, "//div[contains(@class,'col-12 col-md-auto text-center text-start-md')]")
                        return 'warning'
                    except Exception:
                        return 'unknown error'
                        
            except TimeoutException:
                return 'timeout or no result element found'
                
        except Exception as e:
            return f'Error: {str(e)}'
    
    def process_file(self, file_path, output_path):
        """
        Process entire Excel file
        
        Args:
            file_path: Path to input Excel file
            output_path: Path for output Excel file
            
        Returns:
            bool: True if successful, False if failed
        """
        try:
            # Read Excel file
            df = self.read_excel_file(file_path)
            if df is None or len(df) == 0:
                self._log_status("No valid data found in Excel file")
                return False
            
            # Setup browser
            if not self.setup_driver():
                return False
            
            results = []
            total_names = len(df)
            
            self._log_status(f"Starting to process {total_names} names...")
            
            for idx, row in df.iterrows():
                if self.should_stop:
                    self._log_status("Processing stopped by user")
                    break
                
                first_name = row['First Name']
                last_name = row['Last Name']
                gender = row['Gender']
                
                current_name = f"{first_name} {last_name}"
                self._update_progress(idx + 1, total_names, current_name)
                
                # Search registry
                result = self.search_single_name(first_name, last_name, gender)
                
                results.append({
                    'First Name': first_name,
                    'Last Name': last_name,
                    'Gender': gender,
                    'Result': result,
                    'Processed Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                
                # Brief delay between searches
                time.sleep(1)
            
            # Save results to Excel
            results_df = pd.DataFrame(results)
            
            # Create output directory if needed
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Save to Excel with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                results_df.to_excel(writer, sheet_name='Registry Results', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Registry Results']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            self._log_status(f"Results saved to: {output_path}")
            self._log_status(f"Processing complete! Processed {len(results)} names.")
            
            return True
            
        except Exception as e:
            self._log_status(f"Error during processing: {str(e)}")
            return False
        
        finally:
            if self.driver:
                self.driver.quit()
                self.driver = None
    
    def stop_processing(self):
        """Stop the processing"""
        self.should_stop = True
        self._log_status("Stop request received...")
    
    def cleanup(self):
        """Clean up resources"""
        if self.driver:
            self.driver.quit()
            self.driver = None 