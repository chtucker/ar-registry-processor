import pandas as pd
import time
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
import openpyxl
from selenium.common.exceptions import TimeoutException

# Input and output paths
input_path = os.path.join('data', 'extracted_names_v1.csv')
output_path = os.path.join('data', 'registry_results_v1.csv')

# Read the first two valid rows from the new Excel file
wb = openpyxl.load_workbook('data/NEWAdultstotheDatabaseinlastMonth.xlsx')
ws = wb.active
# Assume headers are in row 3, data starts at row 4
rows = []
for i in range(4, 6):  # Rows 4 and 5 (1-indexed)
    first_name = ws[f'B{i}'].value
    last_name = ws[f'D{i}'].value
    gender = ws[f'G{i}'].value.lower() if ws[f'G{i}'].value else 'unknown'
    if first_name and last_name and gender:
        rows.append({'First Name': first_name, 'Last Name': last_name, 'Gender': gender})
names_df = pd.DataFrame(rows)

# Selenium setup
chrome_options = Options()
# Disable location requests
prefs = {"profile.default_content_setting_values.geolocation": 2}
chrome_options.add_experimental_option("prefs", prefs)
# chrome_options.add_argument('--headless')  # Run in headless mode
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-dev-shm-usage')
service = Service('/opt/homebrew/bin/chromedriver')
driver = webdriver.Chrome(service=service, options=chrome_options)

url = 'https://sexoffenderregistry.ar.gov/public/#/'
results = []

for idx, row in names_df.iterrows():
    first = str(row['First Name']) if not pd.isna(row['First Name']) else ''
    last = str(row['Last Name']) if not pd.isna(row['Last Name']) else ''
    driver.get(url)
    try:
        # Wait for page to load
        time.sleep(2)
        # Take a screenshot for debugging
        driver.save_screenshot(f'data/debug_page_{idx}.png')
        # If disclaimer button is present, click it
        try:
            agree_button = WebDriverWait(driver, 3).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(.,'I Agree to the Disclaimer Above')]"))
            )
            # Scroll the main window in increments until the button is enabled
            for _ in range(20):  # Try up to 20 increments
                driver.execute_script("window.scrollBy(0, 300);")
                time.sleep(0.2)
                if agree_button.is_enabled():
                    break
            time.sleep(0.5)
            driver.execute_script("arguments[0].scrollIntoView(true);", agree_button)
            driver.save_screenshot(f'data/debug_disclaimer_{idx}.png')
            agree_button.click()
            time.sleep(1)
        except Exception:
            pass  # Button not present, continue
        # Wait for and click the 'Name Search' tab (button or link)
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//*[contains(text(),'Name Search')]"))
        ).click()
        time.sleep(2)  # Wait for the tab content to load
        driver.save_screenshot(f'data/debug_namesearch_{idx}.png')
        # Wait for input fields
        first_name_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//input[@id='firstName']"))
        )
        last_name_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//input[@id='lastName']"))
        )
        print(f"Found First Name field: {first_name_field}")
        print(f"Found Last Name field: {last_name_field}")
        # Highlight fields for visual confirmation
        driver.execute_script("arguments[0].style.border='3px solid red'", first_name_field)
        driver.execute_script("arguments[0].style.border='3px solid blue'", last_name_field)
        # Fill in first and last name
        first_name_field.clear()
        first_name_field.send_keys(first)
        first_name_field.send_keys(Keys.TAB)
        time.sleep(0.5)
        last_name_field.clear()
        last_name_field.send_keys(last)
        last_name_field.send_keys(Keys.TAB)
        time.sleep(0.5)
        print("Filled in name fields.")
        driver.save_screenshot(f'data/debug_filled_{idx}.png')
        # Set Gender to Male using Select
        gender_value = row['Gender'] if 'Gender' in row and pd.notna(row['Gender']) else 'unknown'
        gender_select = Select(driver.find_element(By.XPATH, "//select[@id='gender']"))
        gender_select.select_by_value(gender_value)
        print(f"Set gender to {gender_value.capitalize()}.")
        # Click Search
        print("Clicking Search button...")
        search_button = driver.find_element(By.XPATH, "//button[@type='submit' and contains(.,'Search')]")
        search_button.click()
        # Take a screenshot after search
        time.sleep(2)
        driver.save_screenshot(f'data/debug_results_{idx}.png')
        # Wait for either the no-result div or the results card
        try:
            # Wait for either the no-result div or the results card
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//div[contains(@class,'no-result') or contains(@class,'col-12 col-md-auto text-center text-start-md')]") )
            )
            # Check for no-result div
            try:
                no_result = driver.find_element(By.XPATH, "//div[contains(@class,'no-result') and contains(.,'No results')]")
                result_text = 'no results found'
            except Exception:
                # If not found, check for results card
                try:
                    result_card = driver.find_element(By.XPATH, "//div[contains(@class,'col-12 col-md-auto text-center text-start-md')]")
                    result_text = 'warning'
                except Exception:
                    result_text = 'unknown error'
        except TimeoutException:
            result_text = 'timeout or no result element found'
        results.append({'First Name': first, 'Last Name': last, 'Gender': gender_value, 'Result': result_text})
    except Exception as e:
        results.append({'First Name': first, 'Last Name': last, 'Gender': gender_value, 'Result': f'Error: {str(e)}'})
    time.sleep(2)  # Delay between searches

driver.quit()

# Save results
results_df = pd.DataFrame(results)
os.makedirs(os.path.dirname(output_path), exist_ok=True)
results_df.to_csv(output_path, index=False)

print(f"Processed {len(results)} names. Results saved to {output_path}") 